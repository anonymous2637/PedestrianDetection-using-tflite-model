import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import os

excel_file = "detections.xlsx"

# Clean creation of a valid Excel file
def create_excel_file():
    df = pd.DataFrame(columns=["Date", "Time", "Person Count"])
    df.to_excel(excel_file, index=False, sheet_name="Detections", engine='openpyxl')

# Ensure Excel file exists and is valid
def ensure_valid_excel():
    try:
        if not os.path.exists(excel_file):
            create_excel_file()
        else:
            # Try opening it to confirm it's valid
            load_workbook(excel_file)
    except Exception as e:
        print(f"Corrupt Excel detected: {e}. Recreating...")
        create_excel_file()

# Check and ensure the Excel file is valid
ensure_valid_excel()

def save_to_excel(person_count):
    ensure_valid_excel()  # Check if the file is valid before writing

    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time_12hr = now.strftime("%I:%M:%S %p")

    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        sheet.append([date, time_12hr, person_count])
        workbook.save(excel_file)
        workbook.close()  # Close to prevent locking issues
    except Exception as e:
        print(f"Error updating Excel file: {e}")
